"""
Seed regulators, vendors, tenants (end users), and their relationships from CSV/XLSX matrices.

Supported files in repo root (optional):
- regulators.csv
- vendors.csv
- end_users.csv (mapped to tenants)
- vendor_regulator_map.csv (columns: vendor, regulator, [notes])
- vendor_enduser_map.csv (columns: vendor, end_user, [notes])
- ksa_regulatory_matrix.csv/.xlsx (columns include standard/regulator/control_id/title/description/version/is_mandatory)
- ksa_connectors_matrix.csv/.xlsx (ignored unless connector model added later)
"""
from __future__ import annotations
import csv
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from app.database import SessionLocal
from app import models

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None  # type: ignore


ROOT = Path(__file__).resolve().parents[1]


def read_csv(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in r.items()})
    return rows


def read_xlsx(path: Path) -> List[Dict[str, Any]]:
    if not openpyxl:
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
    out: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d: Dict[str, Any] = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = val
        out.append(d)
    return out


def upsert_regulators(rows: List[Dict[str, Any]], db) -> int:
    created = 0
    for r in rows:
        name = r.get('name') or r.get('regulator') or r.get('Regulator')
        if not name:
            continue
        o = db.query(models.Regulator).filter(models.Regulator.name == name).first()
        if not o:
            o = models.Regulator(name=str(name), country=str(r.get('country') or 'KSA'), sector=r.get('sector'), website=r.get('website'))
            db.add(o)
            db.commit()
            created += 1
    return created


def upsert_vendors(rows: List[Dict[str, Any]], db) -> int:
    created = 0
    for r in rows:
        name = r.get('name') or r.get('vendor') or r.get('Vendor')
        if not name:
            continue
        o = db.query(models.Vendor).filter(models.Vendor.name == name).first()
        if not o:
            o = models.Vendor(name=str(name), category=r.get('category'), website=r.get('website'), contact_email=r.get('contact_email'), description=r.get('description'))
            db.add(o)
            db.commit()
            created += 1
    return created


def upsert_tenants(rows: List[Dict[str, Any]], db) -> int:
    created = 0
    for r in rows:
        name = r.get('name') or r.get('end_user') or r.get('customer') or r.get('org')
        if not name:
            continue
        o = db.query(models.Tenant).filter(models.Tenant.name == name).first()
        if not o:
            o = models.Tenant(name=str(name))
            db.add(o)
            db.commit()
            created += 1
    return created


def map_vendor_regulator(rows: List[Dict[str, Any]], db) -> int:
    created = 0
    for r in rows:
        vname = r.get('vendor') or r.get('Vendor')
        rname = r.get('regulator') or r.get('Regulator')
        if not vname or not rname:
            continue
        v = db.query(models.Vendor).filter(models.Vendor.name == vname).first()
        g = db.query(models.Regulator).filter(models.Regulator.name == rname).first()
        if not v or not g:
            continue
        exists = db.query(models.VendorRegulator).filter(models.VendorRegulator.vendor_id == v.id, models.VendorRegulator.regulator_id == g.id).first()
        if exists:
            continue
        db.add(models.VendorRegulator(vendor_id=v.id, regulator_id=g.id, notes=r.get('notes')))
        created += 1
    db.commit()
    return created


def map_vendor_tenant(rows: List[Dict[str, Any]], db) -> int:
    created = 0
    for r in rows:
        vname = r.get('vendor') or r.get('Vendor')
        tname = r.get('end_user') or r.get('tenant') or r.get('EndUser')
        if not vname or not tname:
            continue
        v = db.query(models.Vendor).filter(models.Vendor.name == vname).first()
        t = db.query(models.Tenant).filter(models.Tenant.name == tname).first()
        if not v or not t:
            continue
        exists = db.query(models.VendorTenant).filter(models.VendorTenant.vendor_id == v.id, models.VendorTenant.tenant_id == t.id).first()
        if exists:
            continue
        db.add(models.VendorTenant(vendor_id=v.id, tenant_id=t.id, notes=r.get('notes')))
        created += 1
    db.commit()
    return created


def upsert_regulatory_matrix(rows: List[Dict[str, Any]], db) -> Dict[str, int]:
    created_std = 0
    created_ctrl = 0
    for r in rows:
        std_name = r.get('standard') or r.get('Standard') or r.get('framework')
        regulator = r.get('regulator') or r.get('Regulator')
        version = str(r.get('version') or '1.0')
        if not std_name:
            continue
        std = db.query(models.ComplianceStandard).filter(models.ComplianceStandard.name == std_name).first()
        if not std:
            std = models.ComplianceStandard(name=std_name, version=version, description='Imported from matrix')
            if regulator:
                reg = db.query(models.Regulator).filter(models.Regulator.name == regulator).first()
                if not reg:
                    reg = models.Regulator(name=regulator)
                    db.add(reg)
                    db.flush()
                std.regulator_id = reg.id
            db.add(std)
            db.commit()
            db.refresh(std)
            created_std += 1
        cid = r.get('control_id') or r.get('ControlId') or r.get('id') or r.get('number')
        if not cid:
            continue
        exists = db.query(models.Control).filter(models.Control.standard_id == std.id, models.Control.control_id == str(cid)).first()
        if exists:
            continue
        title = r.get('title') or r.get('Title') or 'Untitled'
        desc = r.get('description') or r.get('Description')
        mandatory_val = r.get('is_mandatory')
        try:
            is_mandatory = bool(int(mandatory_val)) if isinstance(mandatory_val, (int, float)) else str(mandatory_val).lower() in {'true','1','yes'}
        except Exception:
            is_mandatory = True
        db.add(models.Control(standard_id=std.id, control_id=str(cid), title=str(title), description=desc, is_mandatory=is_mandatory))
        created_ctrl += 1
    db.commit()
    return {"standards": created_std, "controls": created_ctrl}


def main() -> None:
    db = SessionLocal()
    try:
        # Regulators, vendors, tenants
        for name, fn in (
            ('regulators', 'regulators.csv'),
            ('vendors', 'vendors.csv'),
            ('end_users', 'end_users.csv'),
        ):
            p = ROOT / fn
            if p.exists():
                rows = read_csv(p)
                if name == 'regulators':
                    c = upsert_regulators(rows, db)
                    print(f"regulators: +{c}")
                elif name == 'vendors':
                    c = upsert_vendors(rows, db)
                    print(f"vendors: +{c}")
                elif name == 'end_users':
                    c = upsert_tenants(rows, db)
                    print(f"tenants: +{c}")
        # Mappings
        for fn, handler in (
            ('vendor_regulator_map.csv', map_vendor_regulator),
            ('vendor_enduser_map.csv', map_vendor_tenant),
        ):
            p = ROOT / fn
            if p.exists():
                rows = read_csv(p)
                c = handler(rows, db)
                print(f"{fn}: +{c}")
# Regulatory matrix
        for fn in ('ksa_regulatory_matrix.csv', 'ksa_regulatory_matrix.xlsx'):
            p = ROOT / fn
            if p.exists():
                rows = read_csv(p) if fn.endswith('.csv') else read_xlsx(p)
                res = upsert_regulatory_matrix(rows, db)
                print(f"{fn}: {res}")
        # Connectors matrix placeholder
        for fn in ('ksa_connectors_matrix.csv', 'ksa_connectors_matrix.xlsx'):
            p = ROOT / fn
            if p.exists():
                rows = read_csv(p) if fn.endswith('.csv') else read_xlsx(p)
                created = 0
                for r in rows:
                    name = r.get('name') or r.get('connector') or r.get('Connector')
                    if not name:
                        continue
                    vendor_name = r.get('vendor') or r.get('Vendor')
                    regulator_name = r.get('regulator') or r.get('Regulator')
                    category = r.get('category')
                    desc = r.get('description')
                    # Resolve IDs
                    v = db.query(models.Vendor).filter(models.Vendor.name == vendor_name).first() if vendor_name else None
                    g = db.query(models.Regulator).filter(models.Regulator.name == regulator_name).first() if regulator_name else None
                    # Upsert
                    exists = db.query(models.Connector).filter(models.Connector.name == name, models.Connector.vendor_id == (v.id if v else None), models.Connector.regulator_id == (g.id if g else None)).first()
                    if exists:
                        continue
                    db.add(models.Connector(name=str(name), category=category, vendor_id=(v.id if v else None), regulator_id=(g.id if g else None), description=desc))
                    created += 1
                db.commit()
                print(f"{fn}: connectors +{created}")
    finally:
        db.close()


if __name__ == '__main__':
    main()
